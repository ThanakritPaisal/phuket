#!/usr/bin/env python3
"""
enrich_cbt_google_places.py
============================
Auto-fill the empty columns of LOMA_provider_seed_CBT.xlsx (Providers sheet)
using the Google Places API (New): lat, lng, google_place_id, opening_hours,
contact_phone, price_range, photo_url — for any provider that exists on Google.

WHAT IT DOES (per provider row)
  1. Places "Text Search (New)"  ->  find the business by name + area + "Phuket"
  2. Places "Place Details (New)" ->  pull location, hours, phone, price, photo
  3. Writes the values back into the same xlsx and sets onboard_status = "enriched"
  Rows with no confident match are left blank and flagged (status stays "not started").

SETUP (once)
  1. Create a Google Cloud project:            https://console.cloud.google.com/
  2. Enable "Places API (New)"                 APIs & Services > Library
  3. Create an API key                          APIs & Services > Credentials
  4. Enable billing (free tier is plenty for a pilot; see note below)
  5. pip install requests openpyxl pandas
  6. export GOOGLE_MAPS_API_KEY="your_key_here"   (Windows: set GOOGLE_MAPS_API_KEY=...)

RUN
  python enrich_cbt_google_places.py LOMA_provider_seed_CBT.xlsx

PRICING NOTE (as of 2025 pricing model)
  Text Search = Pro SKU (~5,000 free calls/month).
  Detail fields phone/hours/rating = Enterprise SKU (~1,000 free calls/month).
  76 providers x (1 search + 1 detail) is well within the monthly free tier.

CACHING / ToS
  Google forbids storing most Places content long-term — EXCEPT place_id, which
  you may store forever. Best practice: keep google_place_id, and either (a) let the
  business owner confirm/edit the imported values (that becomes YOUR data), or
  (b) re-fetch live from place_id when needed. Do not silently cache raw fields.

IMPORTANT
  Importing data does NOT prove ownership. Still require OTP / documents / a
  Google-Business-Profile "claim" before marking a provider verified & live.
"""
import os, sys, time, json, requests
import pandas as pd
from openpyxl import load_workbook

API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY")
SEARCH_URL  = "https://places.googleapis.com/v1/places:searchText"
DETAILS_URL = "https://places.googleapis.com/v1/places/{pid}"
PHOTO_URL   = "https://places.googleapis.com/v1/{photo}/media?maxWidthPx=800&key={key}"

SEARCH_FIELDS  = "places.id,places.displayName,places.formattedAddress,places.location"
DETAIL_FIELDS  = ("id,displayName,formattedAddress,location,nationalPhoneNumber,"
                  "regularOpeningHours,priceLevel,rating,userRatingCount,photos")
PRICE_MAP = {"PRICE_LEVEL_INEXPENSIVE":"฿","PRICE_LEVEL_MODERATE":"฿฿",
             "PRICE_LEVEL_EXPENSIVE":"฿฿฿","PRICE_LEVEL_VERY_EXPENSIVE":"฿฿฿฿"}

def text_search(query):
    r = requests.post(SEARCH_URL,
        headers={"Content-Type":"application/json","X-Goog-Api-Key":API_KEY,
                 "X-Goog-FieldMask":SEARCH_FIELDS},
        json={"textQuery":query,"languageCode":"th","regionCode":"TH","maxResultCount":1},
        timeout=20)
    r.raise_for_status()
    places = r.json().get("places",[])
    return places[0] if places else None

def place_details(pid):
    r = requests.get(DETAILS_URL.format(pid=pid),
        headers={"X-Goog-Api-Key":API_KEY,"X-Goog-FieldMask":DETAIL_FIELDS},
        params={"languageCode":"th"}, timeout=20)
    r.raise_for_status()
    return r.json()

def enrich(xlsx_path, sheet="Providers (onboard)"):
    if not API_KEY:
        sys.exit("Set GOOGLE_MAPS_API_KEY first:  export GOOGLE_MAPS_API_KEY='...'")
    df = pd.read_excel(xlsx_path, sheet_name=sheet)
    wb = load_workbook(xlsx_path); ws = wb[sheet]
    header = {c.value: c.column for c in ws[1]}
    def setcell(row, col, val):
        if col in header and val not in (None, ""):
            ws.cell(row=row+2, column=header[col]).value = val

    filled = 0
    for i, r in df.iterrows():
        query = f"{r['name']} {r.get('area','')} Phuket Thailand"
        try:
            hit = text_search(query)
            if not hit:
                print(f"[{i+1}/{len(df)}] no match  · {r['name'][:40]}"); continue
            pid = hit["id"]; d = place_details(pid)
            loc = d.get("location",{})
            setcell(i,"google_place_id", pid)
            setcell(i,"lat", loc.get("latitude"))
            setcell(i,"lng", loc.get("longitude"))
            hrs = d.get("regularOpeningHours",{}).get("weekdayDescriptions")
            if hrs: setcell(i,"opening_hours", " | ".join(hrs))
            setcell(i,"contact_phone", d.get("nationalPhoneNumber"))
            setcell(i,"price_range", PRICE_MAP.get(d.get("priceLevel"),""))
            photos = d.get("photos")
            if photos:
                setcell(i,"photo_url", PHOTO_URL.format(photo=photos[0]["name"], key=API_KEY))
            setcell(i,"onboard_status","enriched")
            filled += 1
            print(f"[{i+1}/{len(df)}] OK       · {r['name'][:40]}  ({loc.get('latitude')},{loc.get('longitude')})")
            time.sleep(0.2)  # be gentle
        except requests.HTTPError as e:
            print(f"[{i+1}/{len(df)}] HTTP {e.response.status_code} · {r['name'][:40]}")
        except Exception as e:
            print(f"[{i+1}/{len(df)}] err {e} · {r['name'][:40]}")

    wb.save(xlsx_path)
    print(f"\nDone. Enriched {filled}/{len(df)} providers. Saved -> {xlsx_path}")
    print("Next: confirm blanks with the community, then set onboard_status = 'live'.")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "LOMA_provider_seed_CBT.xlsx"
    enrich(path)
