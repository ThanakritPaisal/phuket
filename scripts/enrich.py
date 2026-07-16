#!/usr/bin/env python3
"""LOMA provider enrichment from Google Places API (New).
Reads the provider seed xlsx, matches each provider on Google Maps, and writes
an enriched xlsx with the fields the prototype needs + a match-confidence column.
"""
import os, sys, json, time, urllib.request, urllib.parse, urllib.error
from difflib import SequenceMatcher
import openpyxl

KEY = os.environ["GMAPS_KEY"]
SRC = os.environ.get("SRC", "LOMA_provider_seed_CBT.xlsx")
OUT = os.environ.get("OUT", "LOMA_provider_enriched_sample.xlsx")
LIMIT = int(os.environ.get("LIMIT", "5"))
IMG_DIR = os.environ.get("IMG_DIR", "enriched_images")

# Phuket bounding box (island + immediate mainland) for sanity-checking matches
PHK = dict(lat_lo=7.35, lat_hi=8.35, lng_lo=98.15, lng_hi=98.60)
PHK_CENTER = (7.9, 98.35)

DETAIL_FIELDS = ",".join("places." + f for f in [
    "id", "displayName", "formattedAddress", "location", "rating",
    "userRatingCount", "priceLevel", "currentOpeningHours.openNow",
    "regularOpeningHours.weekdayDescriptions", "nationalPhoneNumber",
    "websiteUri", "googleMapsUri", "businessStatus", "primaryTypeDisplayName",
    "editorialSummary", "photos",
    # Deep-link bundle (Places API New): directionsUri / reviewsUri / photosUri /
    # writeAReviewUri / placeUri. The DB write assembles these into the `links` object
    # (see loma-app/src/types.ts ProviderLinks). Google has NO social/booking link field.
    "googleMapsLinks",
])

def post(url, body, field_mask):
    req = urllib.request.Request(
        url, data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json",
                 "X-Goog-Api-Key": KEY,
                 "X-Goog-FieldMask": field_mask}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.load(r)
    except urllib.error.HTTPError as e:
        return {"_error": e.read().decode("utf-8", "replace")}

def norm(s):
    return "".join(ch.lower() for ch in (s or "") if ch.isalnum())

def in_phuket(loc):
    if not loc: return False
    la, ln = loc.get("latitude"), loc.get("longitude")
    return la and PHK["lat_lo"] <= la <= PHK["lat_hi"] and PHK["lng_lo"] <= ln <= PHK["lng_hi"]

def confidence(seed_name, cand):
    if not cand: return "NONE", 0.0
    matched = (cand.get("displayName") or {}).get("text", "")
    ratio = SequenceMatcher(None, norm(seed_name), norm(matched)).ratio()
    loc_ok = in_phuket(cand.get("location"))
    if not loc_ok:               return "LOW (outside Phuket)", ratio
    if ratio >= 0.55:            return "HIGH", ratio
    if ratio >= 0.30:            return "MEDIUM", ratio
    if cand.get("userRatingCount"): return "LOW (name differs)", ratio
    return "LOW (weak match)", ratio

def search(name, area):
    query = f"{name} {area} Phuket".strip()
    body = {
        "textQuery": query,
        "languageCode": "th",
        "maxResultCount": 5,
        "locationBias": {"circle": {
            "center": {"latitude": PHK_CENTER[0], "longitude": PHK_CENTER[1]},
            "radius": 45000.0}},
    }
    res = post("https://places.googleapis.com/v1/places:searchText", body, DETAIL_FIELDS)
    if "_error" in res:
        print("  ! search error:", res["_error"][:300]); return []
    return res.get("places", [])

def photo_url(photo_name, max_px=1000):
    url = (f"https://places.googleapis.com/v1/{photo_name}/media"
           f"?maxHeightPx={max_px}&skipHttpRedirect=true&key={KEY}")
    try:
        with urllib.request.urlopen(url, timeout=30) as r:
            return json.load(r).get("photoUri")
    except Exception as e:
        print("  ! photo error:", e); return None

def download(uri, path):
    try:
        urllib.request.urlretrieve(uri, path); return True
    except Exception as e:
        print("  ! download error:", e); return False

# ---- run ----
os.makedirs(IMG_DIR, exist_ok=True)
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Provider Seed"]
rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]
seed = [dict(zip(hdr, r)) for r in rows[1:] if any(r)]
wb.close()
seed = seed[:LIMIT]

OUT_COLS = ["provider_id", "seed_name", "loma_category", "item_type", "area",
            "match_confidence", "name_similarity", "matched_name", "place_id",
            "matched_address", "latitude", "longitude", "rating",
            "user_rating_count", "price_level", "open_now", "opening_hours",
            "phone", "website", "google_maps_url", "business_status",
            "primary_type", "editorial_summary", "photo_url", "photo_file"]

out_rows = []
for i, s in enumerate(seed, 1):
    name, area = s["provider_name"], s.get("area") or ""
    print(f"[{i}/{len(seed)}] {s['provider_id']}  {name}  ({area})")
    cands = search(name, area)
    # pick best candidate by name similarity, tie-break by review count
    best, best_ratio = None, -1.0
    for c in cands:
        _, r = confidence(name, c)
        score = r + (0.001 if in_phuket(c.get("location")) else -0.5)
        if score > best_ratio:
            best, best_ratio = c, score
    conf, ratio = confidence(name, best)
    row = {k: "" for k in OUT_COLS}
    row.update(provider_id=s["provider_id"], seed_name=name,
               loma_category=s.get("loma_category"), item_type=s.get("item_type"),
               area=area, match_confidence=conf, name_similarity=round(ratio, 2))
    if best:
        loc = best.get("location") or {}
        oh = best.get("regularOpeningHours") or {}
        row.update(
            matched_name=(best.get("displayName") or {}).get("text", ""),
            place_id=best.get("id", ""),
            matched_address=best.get("formattedAddress", ""),
            latitude=loc.get("latitude", ""), longitude=loc.get("longitude", ""),
            rating=best.get("rating", ""), user_rating_count=best.get("userRatingCount", ""),
            price_level=best.get("priceLevel", ""),
            open_now=(best.get("currentOpeningHours") or {}).get("openNow", ""),
            opening_hours=" | ".join(oh.get("weekdayDescriptions", [])),
            phone=best.get("nationalPhoneNumber", ""),
            website=best.get("websiteUri", ""),
            google_maps_url=best.get("googleMapsUri", ""),
            business_status=best.get("businessStatus", ""),
            primary_type=(best.get("primaryTypeDisplayName") or {}).get("text", ""),
            editorial_summary=(best.get("editorialSummary") or {}).get("text", ""),
        )
        photos = best.get("photos") or []
        if photos:
            uri = photo_url(photos[0]["name"])
            if uri:
                row["photo_url"] = uri
                fp = os.path.join(IMG_DIR, f"{s['provider_id']}.jpg")
                if download(uri, fp):
                    row["photo_file"] = fp
        print(f"    -> {conf}: {row['matched_name']}  ({row['rating']}★ x{row['user_rating_count']})")
    else:
        print("    -> NONE (no candidates)")
    out_rows.append(row)
    time.sleep(0.2)

# sort: HIGH first, then MEDIUM, then LOW*, then NONE; tie-break by similarity desc
def rank(c):
    if c.startswith("HIGH"): return 0
    if c.startswith("MEDIUM"): return 1
    if c.startswith("LOW"): return 2
    return 3
out_rows.sort(key=lambda r: (rank(r["match_confidence"]), -float(r["name_similarity"] or 0)))

out = openpyxl.Workbook(); osh = out.active; osh.title = "Enriched"
osh.append(OUT_COLS)
for row in out_rows:
    osh.append([row[k] for k in OUT_COLS])
osh.freeze_panes = "A2"

# summary tab
summ = out.create_sheet("Summary")
from collections import Counter
counts = Counter(r["match_confidence"] for r in out_rows)
summ.append(["match_confidence", "count"])
for k, v in sorted(counts.items(), key=lambda x: rank(x[0])):
    summ.append([k, v])
summ.append(["TOTAL", len(out_rows)])

out.save(OUT)
print("\nSaved:", OUT)
print("Confidence breakdown:", dict(counts))
